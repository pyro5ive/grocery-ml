from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelExportMerger:
    """
    Export multiple pandas DataFrames into a single XLSX file.
    Each DataFrame is written to its own worksheet as a proper Excel Table.
    """

    def __init__(this):
        this.dataframes = []
    ################################################################
    def add_dataframe(this, df: pd.DataFrame, sheetName: str):
        """
        Store DataFrame and sheet name for later export.
        """
        this.dataframes.append((sheetName, df))
    #################################################################
    def write_all(this, output_dir: str, base_name: str):
        """
        Write all stored DataFrames into a single XLSX file.
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        this.outputPath = Path(output_dir) / f"{base_name}_{timestamp}.xlsx"

        for idx, (sheetName, df) in enumerate(this.dataframes):
            mode = "a" if idx > 0 else None
            df.to_excel(
                this.outputPath,
                sheet_name=sheetName,
                index=False,
                mode=mode,
                engine="openpyxl" if mode else None
            )

            workbook = load_workbook(this.outputPath)
            worksheet = workbook[sheetName]

            endRow = worksheet.max_row
            endCol = worksheet.max_column
            endColLetter = worksheet.cell(row=1, column=endCol).column_letter

            tableRef = f"A1:{endColLetter}{endRow}"
            tableName = f"{sheetName}_table"

            table = Table(displayName=tableName, ref=tableRef)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            table.tableStyleInfo = style
            worksheet.add_table(table)
            workbook.save(this.outputPath)
#
    def get_output_path(this) -> str:
        return str(this.outputPath)
#