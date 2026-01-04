import time
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import asyncio
import json
import gc
from sklearn.model_selection import train_test_split
import numpy as np
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
from sklearn.metrics import silhouette_score
#
from school_features import SchoolFeatures
from weather_features import WeatherFeatures
from item_name_utils import ItemNameUtils
from temporal_features_2 import TemporalFeatures
from data_creator import DataCreator
from holiday_features import HolidayFeatures
from wallmart_rcpt_parser import WallmartRecptParser
from winn_dixie_recpt_parser import WinnDixieRecptParser 
from hidden_layer_param_builder import HiddenLayerParamSetBuilder
from weather_service import NwsWeatherService;



class GroceryMLCore:
 
    itemNameUtils = None; 
    weatherService = None; 
    
    def __init__(self):
        pass;
        self.itemNameUtils = ItemNameUtils();
        self.weatherService = NwsWeatherService();
   ###########################################################################################
    def validate_no_empty_columns(self, df):
        # if ANY column has at least one missing value
        print("validate_no_empty_columns()");
        bad_cols = [c for c in df.columns if df[c].isna().any()]
        if bad_cols:
            raise ValueError(f"Columns contain empty values: {bad_cols}")
  ###########################################################################################
 
    def canonicalize(self, df):
    
        patterns = ["prairie-farm-milk","kleinpeter-milk", "kl-milk", "Milk, Fat Free,", "Fat-Free Milk"]
        self.itemNameUtils.canonicalize_items(df, patterns, "milk")
        #
        patterns = ["Bunny Bread", "sandwich-bread", "White Sandwich Bread", "bunny-bread","se-grocers-bread","seg-sandwich-bread", "seg-white-bread"]
        self.itemNameUtils.canonicalize_items(df, patterns, "bread")
        #
        patterns = ["white-bread"]
        self.itemNameUtils.canonicalize_items(df, patterns, "bread")
        #
        patterns = ["blue-bell", "ice-cream", "icescream"]
        self.itemNameUtils.canonicalize_items(df, patterns, "icecream")
        
        patterns = ["dandw-cheese", "kraft-cheese", "se-grocers-cheese", "know-and-love-cheese"]
        self.itemNameUtils.canonicalize_items(df, patterns, "cheese")
        #
        patterns = ["blue-plate-mayo", "blue-plate-mynnase"]
        self.itemNameUtils.canonicalize_items(df, patterns, "mayo")
        #
        patterns = ["gatorade", "powerade", "sports-drink"]
        self.itemNameUtils.canonicalize_items(df, patterns, "gatorade-powerade-sports-drink")
        #
        patterns = [ "tyson","chicken-cutlet", "chicken-leg", "chicken-thigh", "chicken-thighs"]
        self.itemNameUtils.canonicalize_items(df, patterns, "chicken")
        #
        patterns = ["steak","ribs", "pork"]
        self.itemNameUtils.canonicalize_items(df, patterns, "red-meat")
        #
        patterns = ["jimmy-dean",]
        self.itemNameUtils.canonicalize_items(df, patterns, "frozen-breakfast")
        #
        patterns = ["shampoo", "conditioner"]
        self.itemNameUtils.canonicalize_items(df, patterns, "shampoo")     
        #
        patterns = ["soap"]
        self.itemNameUtils.canonicalize_items(df, patterns, "soap")     

        patterns = ["chobani-yogrt-flip", "chobani-yogurt", "yogurt"]
        self.itemNameUtils.canonicalize_items(df, patterns, "yogurt")
        #
        patterns = ["coca-cola", "coca-cola-cola", "cocacola-soda", "coke", "cola"]
        self.itemNameUtils.canonicalize_items(df, patterns, "coke")
        #
        patterns = ["topcare", "top-care"]
        self.itemNameUtils.canonicalize_items(df, patterns, "otcmeds")
           
        patterns = ["little-debbie" , "hugbi-pies", "-hugbi-pies", "candy", "tastykake"]
        self.itemNameUtils.canonicalize_items(df, patterns, "junk-food")
        #
        patterns  = ["cereal", "kellogg-raisn-bran", "kellogg-raisin-bra"]
        self.itemNameUtils.canonicalize_items(df, patterns, "cereal")
        #
        patterns = ["minute-maid-drink", "minute-maid-drinks", "minute-maid-lmnade"]
        self.itemNameUtils.canonicalize_items(df, patterns, "minute-maid-drink")
        #
        patterns = ["egglands-best-egg", "egglands-best-eggs", "eggs"]
        self.itemNameUtils.canonicalize_items(df, patterns, "eggs")
        #
        patterns = ["sprklng-water", "sparkling-ice-wtr", "sparkling-ice", "sparkling-water"]
        self.itemNameUtils.canonicalize_items(df, patterns, "sparkling-ice")
        #
        patterns = ["drinking-water", "purified-drinking",]
        self.itemNameUtils.canonicalize_items(df, patterns, "drinking-water")
        #       
        patterns = ["ground-beef"]
        self.itemNameUtils.canonicalize_items(df, patterns, "ground-beef")
        #
        patterns = ["monster-energy", "monster-enrgy", "monster"]
        self.itemNameUtils.canonicalize_items(df, patterns, "monster-energy")
        #
        patterns = ["smuckers", "jelly"]
        self.itemNameUtils.canonicalize_items(df, patterns, "jelly")
        ### TODO: use nlp libs to remove plural word
        patterns = ["cat-litter", "cats-litter"]
        self.itemNameUtils.canonicalize_items(df, patterns, "cat-litter")
        #
        patterns = ["pizza"]
        self.itemNameUtils.canonicalize_items(df, patterns, "pizza")
        #
        patterns = ["pringles"]
        self.itemNameUtils.canonicalize_items(df, patterns, "pringles")
        
        
        patterns = ["dr-pepper"]
        self.itemNameUtils.canonicalize_items(df, patterns, "dr-pepper")                                      
        #
        patterns = ["aluminum-foil", "foil"]
        self.itemNameUtils.canonicalize_items(df, patterns, "aluminum-foil")                                      
        #
        patterns = ["sour-cream"]
        self.itemNameUtils.canonicalize_items(df, patterns, "sour-cream")
    
        return df;
    ###########################################################################################
    def get_feature_col_names(self, df):
        """Returns feature columns before normalization."""
        return [c for c in df.columns if c.endswith("_feat")]
    ###########################################################################################
    def create_item_supply_level_feat(self, df):
        print("create_item_supply_level_feat()")
    
        try:
            ratio = np.where(
                df["avgDaysBetweenItemPurchases_raw"] > 0,
                df["daysSinceThisItemLastPurchased_raw"] / df["avgDaysBetweenItemPurchases_raw"],
                0.0
            )
    
            df["itemSupplyLevel_feat"] = np.clip(1.0 - ratio, 0.0, 1.0)
    
        except Exception as ex:
            print("create_item_supply_level_feat() failed")
            print(ex)
            raise
    
        return df
    ###########################################################################################
    def create_didBuy_target_col(self, df, colName):
        print(f"creating target col: {colName}");
        df[colName] = 1
        return df; 
    ##############################################################################################
    
    def add_item_total_purchase_count_feat(self, df, feature_name: str):
        """
        Counts total purchases per item (didBuy_target == 1)
        and assigns that count to all rows for the item.
        Throws if any item has zero purchases.
        """
        print("add_item_total_purchase_count_feat()")
    
        item_counts = (
            df.loc[df["didBuy_target"] == 1]
              .groupby("itemId")["itemId"]
              .count()
        )
    
        df[feature_name] = df["itemId"].map(item_counts)
    
        if df[feature_name].isna().any():
            missing_ids = df.loc[df[feature_name].isna(), "itemId"].unique()
            raise ValueError(f"Found itemId(s) with zero purchases: {missing_ids}")
    
        df[feature_name] = df[feature_name].astype(int)
    
        return df
    ##############################################################################################
    
    def build_trip_level_features(self, df):
        print("build_trip_level_features()");
        grouped_df = ( df[["date"]]
            .drop_duplicates()
            .sort_values("date")
            .reset_index(drop=True)
        )
    
        # grouped_df["daysSinceLastTrip_feat"] = TemporalFeatures.create_days_since_last_trip(grouped_df)
        # grouped_df["avgDaysBetweenTrips_feat"] = TemporalFeatures.compute_avg_days_between_trips(grouped_df)
        #
        # grouped_df["daysUntilNextHoliday_feat"]   = grouped_df["date"].apply(HolidayFeatures.compute_days_until_next_holiday)
        # grouped_df["daysSinceLastHoliday_feat"]   = grouped_df["date"].apply(HolidayFeatures.compute_days_since_last_holiday)
        # grouped_df["holidayProximityIndex_feat"]  = grouped_df["date"].apply(HolidayFeatures.compute_holiday_proximity_index)
        #
        grouped_df["daysUntilSchoolStart_feat"]   = grouped_df["date"].apply(SchoolFeatures.compute_days_until_school_start)
        grouped_df["daysUntilSchoolEnd_feat"]     = grouped_df["date"].apply(SchoolFeatures.compute_days_until_school_end)
        grouped_df["schoolSeasonIndex_feat"]      = grouped_df["date"].apply(SchoolFeatures.compute_school_season_index)
        #
        # grouped_df =  TemporalFeatures.add_dst_since_until_features(grouped_df);
        # TemporalFeatures.compute_trip_due_ratio(grouped_df);
        # grouped_df = TemporalFeatures.create_date_features(grouped_df)
        #
        return grouped_df;
    ###########################################################################
    ## TODO:build_purchase_item_freq_cols  is broken 
    # def build_purchase_item_freq_cols(self, df):

    #     print("build_purchase_item_freq_cols()");
    #     freq_windows = [7, 15, 30, 90, 365]
    #     max_w = max(freq_windows)
    
    #     # initialize output columns
    #     for w in freq_windows:
    #         df[f"freq_{w}_feat"] = np.nan
    
    #     # process each itemId independently
    #     result_frames = []
    #     for item_id, group in df.groupby("itemId", group_keys=False):
    #         group = group.copy()
    #         group = group.sort_values("date").reset_index(drop=True)
    #         history = []
    
    #         col_date = group.columns.get_loc("date")
    #         col_buy = group.columns.get_loc("didBuy_target")
    #         col_freq = {w: group.columns.get_loc(f"freq_{w}_feat") for w in freq_windows}
    
    #         for i in range(len(group)):
    #             cur_date = group.iat[i, col_date]
    
    #             # record a purchase occurrence
    #             if group.iat[i, col_buy] == 1:
    #                 history.append(cur_date)
    
    #             # prune old entries once using max window
    #             cutoff_max = cur_date - pd.Timedelta(days=max_w)
    #             history = [d for d in history if d >= cutoff_max]
    
    #             # compute window counts
    #             for w in freq_windows:
    #                 cutoff = cur_date - pd.Timedelta(days=w)
    #                 count = 0
    #                 for d in history:
    #                     if d >= cutoff:
    #                         count += 1
    #                 group.iat[i, col_freq[w]] = count
    
    #         result_frames.append(group)
    
    #     return pd.concat(result_frames, ignore_index=True)
    ###########################################################################################    
    def insert_negative_samples(self, df):
        print("insert_negative_samples()")
        
        # keep a lookup of itemId -> item name (each itemId maps to exactly one name)
        item_lookup = (
            df[["itemId", "item"]]
            .drop_duplicates(subset=["itemId"])
        )

        targetColName = "didBuy_target";
        df["didBuy_target"] = 1

        # 2. full grid
        all_items = df["itemId"].unique()
        all_dates = df["date"].unique()
        full = (
            pd.MultiIndex.from_product([all_dates, all_items], names=["date", "itemId"])
            .to_frame(index=False)
        )
    
    
        df_full = full.merge(df,  on=["date", "itemId"], how="left")
        
        # 4. fill missing didBuy
        df_full["didBuy_target"] = df_full["didBuy_target"].fillna(0).astype(int)
    
        # 5. fill missing item names using lookup
        df_full = df_full.merge(item_lookup, on="itemId", how="left", suffixes=("", "_lookup"))
        df_full["item"] = df_full["item"].fillna(df_full["item_lookup"])
        df_full = df_full.drop(columns=["item_lookup"])

        # fill mssing source cols: 
        df_full["source"] = df_full["source"].fillna("_neg_sample_").astype(str)
        
        return df_full.copy()
    ###########################################################################################
    def build_winn_dixie_additional_text_rcpts_df(self, folderPath):
        recptParser = WinnDixieRecptParser()
        rows = []
        for p in Path(folderPath).glob("*.txt"):
            result = recptParser.parse(p.read_text(encoding="utf-8", errors="ignore"))
            for r in result["items"]:
                rows.append({
                    "source": p.name,
                    "date": result["date"],
                    "time": result["time"],
                    #"manager": result["manager"],
                    #"cashier": result["cashier"],
                    "item": r["item"]
                    #"qty": r["qty"],
                    #"reg": r["reg"],
                    #"youPay": r["youPay"],
                    #"reportedItemsSold": result["reported"],
                    #"rowsMatchReported": result["validation"]["rowsMatchReported"],
                    #"qtyMatchReported": result["validation"]["qtyMatchReported"],
                })
    
        additional_rcpts_df = pd.DataFrame(rows)
        
        additional_rcpts_df["date"] = pd.to_datetime(additional_rcpts_df["date"])
        additional_rcpts_df["time"] = additional_rcpts_df["time"].astype(str)
        
        additional_rcpts_df = WinnDixieRecptParser.remove_duplicate_receipt_files(additional_rcpts_df)
        
        #additional_rcpts_df["item"] = additional_rcpts_df["item"].str.replace(r"^know-and-love\s*", "", regex=True, case=False).str.strip()
        additional_rcpts_df["item"] = additional_rcpts_df["item"].str.replace(r"^seg\s*", "", regex=True, case=False).str.strip()
        additional_rcpts_df["item"] = additional_rcpts_df["item"].str.replace(r"^kandl\s*", "", regex=True, case=False).str.strip()

        additional_rcpts_df = additional_rcpts_df.sort_values(by=["date", "time"]).reset_index(drop=True)
        additional_rcpts_df = additional_rcpts_df.drop(columns=["time"])
        return additional_rcpts_df;
    ###########################################################################################
    def build_winn_dixie_df(self, path):
        recptParser = WinnDixieRecptParser()
        rows = []
        for p in Path(path).glob("*.txt"):
            result = recptParser.parse(p.read_text(encoding="utf-8", errors="ignore"))
            for r in result["items"]:
                rows.append({
                    "source": p.name,
                    "date": result["date"],
                    "time": result["time"],
                    #"manager": result["manager"],
                    #"cashier": result["cashier"],
                    "item": r["item"]
                    #"qty": r["qty"],
                    #"reg": r["reg"],
                    #"youPay": r["youPay"],
                    #"reportedItemsSold": result["reported"],
                    #"rowsMatchReported": result["validation"]["rowsMatchReported"],
                    #"qtyMatchReported": result["validation"]["qtyMatchReported"],
                })
    
        winndixie_df = pd.DataFrame(rows)
        
        winndixie_df["date"] = pd.to_datetime(winndixie_df["date"])
        winndixie_df["time"] = winndixie_df["time"].astype(str)
        
        winndixie_df = WinnDixieRecptParser.remove_duplicate_receipt_files(winndixie_df)
        
        winndixie_df["item"] = winndixie_df["item"].str.replace(r"^know-and-love\s*", "", regex=True, case=False).str.strip()
        winndixie_df["item"] = winndixie_df["item"].str.replace(r"^seg\s*", "", regex=True, case=False).str.strip()
        winndixie_df["item"] = winndixie_df["item"].str.replace(r"^kandl\s*", "", regex=True, case=False).str.strip()
        winndixie_df["source"] = "winndixie";
        winndixie_df = winndixie_df.sort_values(by=["date", "time"]).reset_index(drop=True)
        winndixie_df = winndixie_df.drop(columns=["time"])
        return winndixie_df;
    ###########################################################################################
    def export_df_to_excel_table(self, df, base_path_without_ext, sheet_name="Data"):
        """
        Export a pandas DataFrame to an Excel file as a proper Excel Table
        with no duplicated header rows.
        """
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo

        file_path = f"{base_path_without_ext}.xlsx"
        print(f"Writing XLSX: {file_path}")

        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
    
        end_row = worksheet.max_row
        end_col = worksheet.max_column
        end_col_letter = worksheet.cell(row=1, column=end_col).column_letter
    
        table_ref = f"A1:{end_col_letter}{end_row}"
        table = Table(displayName="DataTable", ref=table_ref)
    
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    
        table.tableStyleInfo = style
        worksheet.add_table(table)
    
        workbook.save(file_path)
        print(f"   XLSX Done: {file_path}")
    ###########################################################################################    
    def export_dataframe_to_csv(self, df, base_path_without_ext):
        file_path = f"{base_path_without_ext}.csv"
        print(f"Writing CSV: {file_path}")
        df.to_csv(file_path, index=True)
        print(f"  CSV done: {file_path}")
    ###########################################################################################    
    def export_dataframe_to_parquet(self, df, base_path_without_ext):
        file_path = f"{base_path_without_ext}.parquet"
        print(f"Writing PARQUET: {file_path}")
        df.to_parquet(file_path, index=True)
        print(f"  PARQUET done: {file_path}")
    ###########################################################################################
    def export_dataframes_with_exp_name(self, dataframes, path):
        for name, df in dataframes.items():
            base = os.path.join(path, f"{name}")
            self.export_df_to_excel_table(df, base, sheet_name=f"{name}")
            #self.export_dataframe_to_csv(df, base)
    ###########################################################################################    
    def write_json(self, obj, path):
        f = open(path, "w")
        json.dump(obj, f, indent=2)
        f.close()  
    ###########################################################################################    
