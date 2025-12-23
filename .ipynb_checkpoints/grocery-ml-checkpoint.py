

import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import os
from datetime import datetime
import asyncio
import json

import gc
import tensorflow as tf
from tensorflow.keras import layers, models

from sklearn.model_selection import train_test_split
import numpy as np
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
from sklearn.metrics import silhouette_score

#from dataset_utils import DatasetUtils
from temporal_features import TemporalFeatures
from holiday_features import HolidayFeatures
from wallmart_rcpt_parser import WallmartRecptParser
from winn_dixie_recpt_parser import WinnDixieRecptParser 
from hidden_layer_param_builder import HiddenLayerParamSetBuilder
asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


class GroceryML:

    def __init__(self):
        self.item_to_id = None
        self.id_to_item = None
        
    ###########################################################################################
    def BuildCombinedDataset(self):

        winndixie_df = self.BuildWinnDixie();
        wallmaert_df = self.BuildWallMart();
        weather_df = self.BuildWeather();
        
        self.combined_df = pd.concat([winndixie_df, wallmaert_df], ignore_index=True)
        self.combined_df["item"] = (self.combined_df["item"]
                .str.replace(r"^\s*[-–—]\s*", "", regex=True)
                .str.strip()
        )
        self.CreateItemIds();
        self.Canonicalize();
        
        self.combined_df = self.combined_df.merge(weather_df, on="date", how="left")
                
        self.BuildTripLevelFeatures();
        
        self.combined_df["freq7_over30"], self.combined_df["freq30_over365"]  
            = TemporalFeatures.compute_freq_ratios(self.combined_df["freq_7"], self.combined_df["freq_30"], self.combined_df["freq_365"])

        #     combined_df = pd.concat(
        #     [winndixie_df, wallmart_df[["date", "item", "source"]]],
        #     ignore_index=True
        # )
        self.combined_df["item_due_ratio"] = compute_due_ratio(combined_df)
        # remove - 
             
        freq_windows = [7, 15, 30, 90, 365]
        max_w = max(freq_windows)
        # initialize columns
        for w in freq_windows:
            self.combined_df[f"freq_{w}_feat"] = np.nan
        self.combined_df = (self.combined_df.groupby("itemId", group_keys=False).apply(fill_freq))
        # ============================================================
        # MERGE HABIT FEATURES
        # ============================================================
        habit_df = build_habit_features(combined_df)
        
        self.combined_df = self.combined_df.merge(habit_df, on="itemId",how="left")
        
        self.combined_df[["habitFrequency_feat", "habitSpan_feat", "habitDecay_feat"]] = (
            self.combined_df[["habitFrequency_feat", "habitSpan_feat", "habitDecay_feat"]].fillna(0.0)
        ) 
    ###########################################################################################

    def CreateItemIds(self, df):
        if self.id_to_item is not None:
            raise RuntimeError("ItemId mapping already initialized")

        unique_items = sorted(df["item"].unique())
        self.item_to_id = {item: idx for idx, item in enumerate(unique_items)}
        self.id_to_item = {idx: item for item, idx in self.item_to_id.items()}

        df["itemId"] = df["item"].map(self.item_to_id)
        df.reset_index(drop=True, inplace=True)
        return df
    ###########################################################################################
    
    def MapItemIdsToNames(self, df, col_name="item"):
        if self.id_to_item is None:
            raise RuntimeError("ItemId mapping not initialized")

        df[col_name] = df["itemId"].map(self.id_to_item)
        return df
    ###########################################################################################

    def GetIdToItem(self):
        if self.id_to_item is None:
            raise RuntimeError("ItemId mapping not initialized")
        return self.id_to_item    
    ###########################################################################################

    def build_habit_features(self,df, tau_days=120):
        df = df.copy()
        df["date"] = pd.to_datetime(df["date"])
    
        total_trips = df["date"].nunique()
        timeline_days = (df["date"].max() - df["date"].min()).days or 1
    
        rows = []
    
        for itemId, g in df.groupby("itemId"):
            buys = g[g["didBuy_target"] == 1]["date"]
    
            if len(buys) == 0:
                rows.append({
                    "itemId": itemId,
                    "habitFrequency_feat": 0.0,
                    "habitSpan_feat": 0.0,
                    "habitDecay_feat": 0.0,
                })
                continue
    
            first = buys.min()
            last = buys.max()
    
            habitFrequency = len(buys) / total_trips
            habitSpan = (last - first).days / timeline_days
            days_since_last = (df["date"].max() - last).days
            habitDecay = np.exp(-days_since_last / tau_days)
    
            rows.append({
                "itemId": itemId,
                "habitFrequency_feat": habitFrequency,
                "habitSpan_feat": habitSpan,
                "habitDecay_feat": habitDecay,
            })
    
        return pd.DataFrame(rows)
    ###########################################################################################
    
    def compute_due_ratio(df, cap=3.0):
        ratio = df["daysSinceLastPurchase_feat"] / df["avgDaysBetweenPurchases_feat"]
        ratio = ratio.replace([np.inf, -np.inf], np.nan).fillna(0)
        return ratio.clip(0, cap)
    ###########################################################################################

    def compute_due_score(df,itemId=None,use_sigmoid=True,normalize=False, weights=None):
        
        if weights is None:
            weights = {
                "daysSinceLastPurchase_feat": 1.5,
                "freq_30_feat": 1.0,
                "freq_90_feat": 0.5
            }
    
        # --------------------------------------------------------
        # Optional itemId filter
        # --------------------------------------------------------
        if itemId is not None:
            df = df[df["itemId"] == itemId].copy()
        else:
            df = df.copy()
    
        # --------------------------------------------------------
        # RAW linear score (pre-normalization)
        # --------------------------------------------------------
        df["due_score_raw_feat"] = (
            weights["daysSinceLastPurchase_feat"] * df["daysSinceLastPurchase_feat"]
          + weights["freq_30_feat"]              * df["freq_30_feat"]
          + weights["freq_90_feat"]              * df["freq_90_feat"]
        )
    
        # --------------------------------------------------------
        # Final due_score
        # --------------------------------------------------------
        if use_sigmoid:
            df["due_score"] = 1 / (1 + np.exp(-df["due_score_raw"]))
    
        elif normalize:
            mean = df["due_score_raw"].mean()
            std  = df["due_score_raw"].std() or 1.0
            df["due_score"] = (df["due_score_raw"] - mean) / std
    
        else:
            df["due_score"] = df["due_score_raw"]
    
        return df
    ###########################################################################################

    def BuildTripLevelFeatures(self):
        # 1. Build grouped table (one row per trip date)

        grouped = ( self.combined_df[["date"]]
            .drop_duplicates()
            .sort_values("date")
            .reset_index(drop=True)
        )      
        
        grouped["daysSinceLastTrip_feat"] = TemporalFeatures.ComputeDaysSinceLastTrip(grouped)
        grouped["avgDaysBetweenTrips_feat"] = TemporalFeatures.ComputeAvgDaysBetweenTrips(grouped)
        
        # 3. Holiday / School features
        grouped["daysUntilNextHoliday_feat"] = grouped["date"].apply(HolidayFeatures.ComputeDaysUntilNextHoliday)
        grouped["daysSinceLastHoliday_feat"] = grouped["date"].apply(HolidayFeatures.ComputeDaysSinceLastHoliday)
        grouped["holidayProximityIndex_feat"] = grouped["date"].apply(HolidayFeatures.ComputeHolidayProximityIndex)
        grouped["daysUntilSchoolStart_feat"] = grouped["date"].apply(HolidayFeatures.ComputeDaysUntilSchoolStart)
        grouped["daysUntilSchoolEnd_feat"]   = grouped["date"].apply(HolidayFeatures.ComputeDaysUntilSchoolEnd)
        grouped["schoolSeasonIndex_feat"]    = grouped["date"].apply(HolidayFeatures.ComputeSchoolSeasonIndex)
               
        grouped = TemporalFeatures.CreateDateFeatures(grouped)
        
        # merge in weather
        # grouped = grouped.merge(df_weather, on="date", how="left")
        
        self.combined_df = self.combined_df.merge(grouped, on="date", how="left")
    ###########################################################################################

    def BackFillitems(self):
        # 1. Mark actual purchases in the raw receipt rows
        self.combined_df["didBuy"] = 1
        # 2. Build complete grid
        all_items = self.combined_df["itemId"].unique()
        all_dates = self.combined_df["date"].unique()
        
        full = (
            pd.MultiIndex.from_product(
                [all_dates, all_items], 
                names=["date", "itemId"]
            ).to_frame(index=False)
        )
        
        # 3. Merge raw purchases onto the full grid
        df_full = full.merge(
            self.combined_df[["date", "itemId", "item", "source", "didBuy"]],
            on=["date", "itemId"],
            how="left"
        )
        
        # 4. Fill missing purchases with didBuy=0
        df_full["didBuy"] = df_full["didBuy"].fillna(0).astype(int)
        
        # 5. NOW REPLACE combined_df with df_full
        self.combined_df = df_full.copy()
    ###########################################################################################

    def Canonicalize(self):
        milk_patterns = ["know-and-love-milk", "kandl-milk", "prairie-farm-milk","kleinpeter-milk", "kl-milk", "Milk, Fat Free,", "Fat-Free Milk"]
        self.CanonicalizeItems(self.combined_df, milk_patterns, "milk")
        #
        bread_patterns = ["bunny-bread","se-grocers-bread","seg-sandwich-bread", "seg-white-bread"]
        self.CanonicalizeItems(self.combined_df, bread_patterns, "bread")
        #
        cheese_patterns = ["dandw-cheese", "kraft-cheese", "se-grocers-cheese", "know-and-love-cheese"]
        self.CanonicalizeItems(self.combined_df, cheese_patterns, "cheese")
        #
        mayo_patterns = ["blue-plate-mayo", "blue-plate-mynnase"]
        self.CanonicalizeItems(self.combined_df, mayo_patterns, "mayo")
        #
        chicken_patterns = ["chicken-cutlet", "chicken-leg", "chicken-thigh", "chicken-thighs"]
        self.CanonicalizeItems(self.combined_df, chicken_patterns, "chicken")
        #
        yogurt_patterns = ["chobani-yogrt-flip", "chobani-yogurt"]
        self.CanonicalizeItems(self.combined_df, yogurt_patterns, "yogurt")
        #
        coke_patterns = ["coca-cola", "coca-cola-cola", "cocacola-soda", "coke", "cola"]
        self.CanonicalizeItems(self.combined_df, coke_patterns, "coke")
        #
        hugbi_patterns = ["hugbi-pies", "-hugbi-pies"]
        self.CanonicalizeItems(self.combined_df, hugbi_patterns, "hugbi-pies")
        #
        ceralPaterns  = ["cereal"]
        self.CanonicalizeItems(self.combined_df, ceralPaterns, "ceral")
        #
        minute_maid_patterns = ["minute-maid-drink", "minute-maid-drinks", "minute-maid-lmnade"]
        self.CanonicalizeItems(self.combined_df, minute_maid_patterns, "minute-maid-drink")
        #
        eggs_pattern = ["egglands-best-egg", "egglands-best-eggs", "eggs"]
        self.CanonicalizeItems(self.combined_df, eggs_pattern, "eggs")
    ###########################################################################################
    def BuildWallMart(self):
        wallmart_raw = WallmartRecptParser.ImportWallMart("./walmart")
        ## rename cols
        wallmart_df = wallmart_raw[["Order Date","Product Description", "source"]].copy()
        wallmart_df = wallmart_df.rename(columns={
            "Order Date": "date",
            "Product Description": "item"
        })
        
        wallmart_df["date"] = pd.to_datetime(wallmart_df["date"])
        return wallmart_df;
    ###########################################################################################
    def BuildWinnDixie(self):
        recptParser  = WinnDixieRecptParser();
            for p in Path("winndixie rcpts/StevePhone2/pdf/text").glob("*.txt"):
            result = recptParser.parse(p.read_text(encoding="utf-8", errors="ignore"))
            for r in result["items"]:
                rows.append({
                    "source": p.name,
                    "date": result["date"],
                    "time": result["time"],
                    #"manager": result["manager"],
                    #"cashier": result["cashier"],
                    "item": r["item"]
                    #"qty": r["qty"],
                    #"reg": r["reg"],
                    #"youPay": r["youPay"],
                    #"reportedItemsSold": result["reported"],
                    #"rowsMatchReported": result["validation"]["rowsMatchReported"],
                    #"qtyMatchReported": result["validation"]["qtyMatchReported"],
                })
    
        winndixie_df = pd.DataFrame(rows)
        
        winndixie_df["date"] = pd.to_datetime(winndixie_df["date"])
        winndixie_df["time"] = winndixie_df["time"].astype(str)
        
        winndixie_df = WinnDixieRecptParser.remove_duplicate_receipt_files(winndixie_df)
        
        winndixie_df = winndixie_df.sort_values(by=["date", "time"]).reset_index(drop=True)
        winndixie_df = winndixie_df.drop(columns=["time"])
        return winndixie_df;
    ###########################################################################################

    def BuildWeather(self):
        # --- WEATHER PREP ---
        weatherCols=["datetime", "temp", "humidity", "feelslike", "dew", "precip"]
        df_weather = pd.read_csv("datasets/VisualCrossing-70062 2000-01-01 to 2025-12-14.csv", usecols=weatherCols)
        df_weather["datetime"] = pd.to_datetime(df_weather["datetime"])
        df_weather = df_weather.set_index("datetime").sort_index()

        df_weather["temp_5day_avg_feat"] = df_weather["temp"].rolling(5, min_periods=1).mean()
        df_weather["feelsLike_5day_avg_feat"] = df_weather["feelslike"].rolling(5, min_periods=1).mean()
        df_weather["dew_5day_avg_feat"] = df_weather["dew"].rolling(5, min_periods=1).mean()
        df_weather["humidity_5day_avg_feat"] = df_weather["humidity"].rolling(5, min_periods=1).mean()
        df_weather["precip_5day_avg_feat"] = df_weather["precip"].rolling(5, min_periods=1).mean()

        df_weather = df_weather.drop(columns=["temp", "humidity", "feelslike", "dew", "precip"])

        # convert index to date for merging
        df_weather["date"] = df_weather.index.date
        df_weather["date"] = pd.to_datetime(df_weather["date"])
        df_weather = df_weather.set_index("date")
        return df_weather;
    ###########################################################################################

    def export_df_to_excel_table(df, file_path, sheet_name="Data"):
        """
        Export a pandas DataFrame to an Excel file as a proper Excel Table
        with no duplicated header rows.
        """
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo
    
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
    
        end_row = worksheet.max_row
        end_col = worksheet.max_column
        end_col_letter = worksheet.cell(row=1, column=end_col).column_letter
    
        table_ref = f"A1:{end_col_letter}{end_row}"
        table = Table(displayName="DataTable", ref=table_ref)
    
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    
        table.tableStyleInfo = style
        worksheet.add_table(table)
    
        workbook.save(file_path)
    ###########################################################################################
    
    def CanonicalizeItems(self, df, patterns, canonical_name):
        """
        For each pattern in `patterns`, find rows where `item` contains the pattern
        and replace df['item'] with `canonical_name`.
        """
        for p in patterns:
            mask = df["item"].str.contains(p, case=False, na=False)
            df.loc[mask, "item"] = canonical_name    
    ###########################################################################################
    
    def export_df(dataframes, dir):
        for name, df in dataframes.items():
            csv_path = os.path.join(dir, f"{name}.csv")
            xlsxPath = os.path.join(dir, f"{name}.xlsx")
            print(f"Writing CSV: {csv_path}")
            df.to_csv(csv_path, index=True)
            print(f"Writing XLSX: {xlsxPath}")
            export_df_to_excel_table(df, xlsxPath, sheet_name=f"{name}")
    ###########################################################################################
      
    def save_experiment(model, history, dataframes, build_params, train_params, norm_params, base_dir):
        name_parts = []
        if "embedding_dim" in build_params:
            name_parts.append(f"emb{build_params['embedding_dim']}")
        if "layers" in build_params:
            hl = "-".join(str(x) for x in build_params["layers"])
            name_parts.append(f"hl{hl}")
        if "epochs" in train_params:
            name_parts.append(f"ep{train_params['epochs']}")
        if "output_activation" in build_params:
            name_parts.append(f"outAct_{build_params['output_activation']}")
    
        exp_name = "__".join(name_parts) if name_parts else "exp_unlabeled"
        exp_dir = os.path.join(base_dir, exp_name)
        print("Saving Exp: ", exp_dir)
        
        os.makedirs(exp_dir, exist_ok=True)
    
        export_df(dataframes, exp_dir)
    
        model.save(os.path.join(exp_dir, "model"))
        model.save_weights(os.path.join(exp_dir, "weights.h5"))
    
        history_path = os.path.join(exp_dir, "history.json")
        history_file = open(history_path, "w")
        json.dump(history.history, history_file, indent=2)
        history_file.close()
    
        build_params_path = os.path.join(exp_dir, "build_params.json")
        build_params_file = open(build_params_path, "w")
        json.dump(build_params, build_params_file, indent=2)
        build_params_file.close()
    
        train_params_path = os.path.join(exp_dir, "train_params.json")
        train_params_file = open(train_params_path, "w")
        json.dump(train_params, train_params_file, indent=2)
        train_params_file.close()
    
        norm_params_path = os.path.join(exp_dir, "norm_params.json")
        norm_params_file = open(norm_params_path, "w")
        json.dump(norm_params, norm_params_file, indent=2)
        norm_params_file.close()
    
        print("Saved experiment →", exp_dir)

    ###############################################
    
    def fit_normalization_params(combined_df):
        params = {}
        feature_cols = [c for c in combined_df.columns if c.endswith("_feat")]
        cyc_cols = [c for c in feature_cols if c.endswith("_cyc_feat")]
        num_cols = [c for c in feature_cols if c not in cyc_cols]
    
        for col in num_cols:
            params[col] = {
                "mean": combined_df[col].mean(),
                "std": combined_df[col].std()
            }
    
        for col in cyc_cols:
            params[col] = {
                "period": TemporalFeatures.get_period_for_column(col)
            }
    
        return params
    ###############################################
    
    def normalize_features(combined_df, norm_params):
    
        print("Normalizin df");
        normalized_df = combined_df.copy()
        for col, cfg in norm_params.items():
            if col.endswith("_cyc_feat"):
                sin_col, cos_col = TemporalFeatures.encode_sin_cos(combined_df[col], cfg["period"])
                normalized_df[f"{col}_sin_norm"] = sin_col
                normalized_df[f"{col}_cos_norm"] = cos_col
                normalized_df.drop(columns=[col], inplace=True)
    
            else:
                mean_val = cfg["mean"]
                std_val = cfg["std"]
                norm_col = col.replace("_feat", "_norm")
    
                if std_val == 0:
                    normalized_df[norm_col] = 0.0
                else:
                    normalized_df[norm_col] = (combined_df[col] - mean_val) / std_val
    
                normalized_df.drop(columns=[col], inplace=True)
    
        return normalized_df
    ###############################################

    def build_and_compile_model(feat_cols_count, item_count, build_params):
        num_in = layers.Input(shape=(feat_cols_count,))
        item_in = layers.Input(shape=(), dtype="int32")
    
        emb = layers.Embedding(
            input_dim=item_count,
            output_dim=build_params["embedding_dim"]
        )(item_in)
    
        x = layers.Concatenate()([num_in, layers.Flatten()(emb)])
    
        for neuron_count in build_params["layers"]:
            x = layers.Dense(neuron_count, activation=build_params["activation"])(x)
    
        out = layers.Dense(1, activation=build_params["output_activation"])(x)
    
        model = models.Model(inputs=[num_in, item_in], outputs=out)
    
        optimizer_name = build_params.get("optimizer", "adam")
        learning_rate = build_params.get("learning_rate")
    
        if optimizer_name == "adam":
            optimizer = tf.keras.optimizers.Adam(learning_rate=learning_rate)
        elif optimizer_name == "adamw":
            optimizer = tf.keras.optimizers.AdamW(learning_rate=learning_rate)
        else:
            raise ValueError(f"Unsupported optimizer: {optimizer_name}")
    
        model.compile(
            optimizer=optimizer,
            loss=build_params.get("loss", "mse"),
            metrics=build_params.get("metrics", ["mae"])
        )
    
        return model
    ###############################################


    def train_model(model, df, feature_cols, target_col, train_params):
        x_feat = df[feature_cols].to_numpy(np.float32)
        x_item = df["itemId"].to_numpy(np.int32)
        y = df[target_col].to_numpy(np.float32)
    
        x_feat_tr, x_feat_te, x_item_tr, x_item_te, y_tr, y_te = train_test_split(
            x_feat, x_item, y, test_size=0.2, random_state=42
        )
    
        history = model.fit(
            [x_feat_tr, x_item_tr],
            y_tr,
            validation_split=0.1,
            epochs=train_params["epochs"],
            batch_size=train_params.get("batch_size", 32),
            verbose=1
        )
    
        return history
    ###############################################


    def build_prediction_input(combined_df, prediction_date, norm_params):
    
        print("Building Prediction df");
        print(f"Prediction date: {prediction_date}")
        
        latest_rows_df = (combined_df.sort_values("date").groupby("itemId").tail(1).copy())
        latest_rows_df["date"] = prediction_date       
        latest_rows_df["daysSinceLastTrip_feat"] = (prediction_date - combined_df["date"].max()).days
        latest_rows_df["avgDaysBetweenTrips_feat"] = combined_df["avgDaysBetweenTrips_feat"].iloc[-1]
    
        latest_rows_df["daysUntilNextHoliday_feat"] = HolidayFeatures.ComputeDaysUntilNextHoliday(prediction_date)
        latest_rows_df["daysSinceLastHoliday_feat"] = HolidayFeatures.ComputeDaysSinceLastHoliday(prediction_date)
        latest_rows_df["holidayProximityIndex_feat"] = HolidayFeatures.ComputeHolidayProximityIndex(prediction_date)
        latest_rows_df["daysUntilSchoolStart_feat"] = SchoolFeatures.ComputeDaysUntilSchoolStart(prediction_date)
        latest_rows_df["daysUntilSchoolEnd_feat"] = SchoolFeatures.ComputeDaysUntilSchoolEnd(prediction_date)
        latest_rows_df["schoolSeasonIndex_feat"] = SchoolFeatures.ComputeSchoolSeasonIndex(prediction_date)
    
        latest_rows_df["year_feat"] = prediction_date.year
        latest_rows_df["month_cyc_feat"] = prediction_date.month
        latest_rows_df["day_cyc_feat"] = prediction_date.day
        latest_rows_df["dow_cyc_feat"] = prediction_date.weekday()
        latest_rows_df["doy_feat"] = prediction_date.timetuple().tm_yday
        latest_rows_df["quarter_feat"] = ((prediction_date.month - 1) // 3) + 1
    
        for item_id in latest_rows_df["itemId"].values:
            hist = combined_df[combined_df["itemId"] == item_id]
    
            #FeatureBuilders.compute_frequency_features(hist, latest_rows_df, item_id, prediction_date)
    
            #FeatureBuilders.compute_habit_features(hist, latest_rows_df, item_id, prediction_date)
    
        if "didBuy_target" in latest_rows.columns:
            latest_rows_df.drop(columns=["didBuy_target"], inplace=True)
    
        export_df_to_excel_table(latest_rows_df, "latest_rows_df.xlsx", ".");
        
        normalized_latest_rows_df = normalize_features(latest_rows_df, norm_params)
    
        feature_cols = [c for c in normalized_latest_rows_df.columns if c.endswith("_norm")]
    
        x_features = normalized_latest_rows_df[feature_cols].to_numpy(np.float32)
        x_item_idx = normalized_latest_rows_df["itemId"].to_numpy(np.int32)
        export_df_to_excel_table(normalized_latest_rows_df, "normalized_latest_rows_df.xlsx", ".");
        normalized_latest_rows_df = self.MapItemIdsToNames();
        
        return {
            "prediction_df": normalized_latest_rows_df,
            "x_features": x_features,
            "x_item_idx": x_item_idx,
            "feature_cols": feature_cols
        }
    ###############################################
    
    
    def RunExperiment(combined_df, modelBuildParams, modelTrainParams, baseDir):
        norm_params = fit_normalization_params(combined_df)
        normalized_df = normalize_features(combined_df, norm_params)
    
        feature_cols = [c for c in normalized_df.columns if c.endswith("_norm")]
        target_cols = [c for c in normalized_df.columns if c.endswith("_target")]
    
        if len(target_cols) != 1:
            raise ValueError("Exactly one target column is required")
    
        target_col = target_cols[0]
    
        feat_cols_count = len(feature_cols)
        item_count = int(normalized_df["itemId"].max()) + 1
    
        model = build_and_compile_model(feat_cols_count, item_count, modelBuildParams)
    
        history = train_model(model, normalized_df, feature_cols, target_col, modelTrainParams)
       
        # pred_input = build_prediction_input_df(combined_df, normalized_df["date"].max(), norm_params)
        pred_input = build_prediction_input(combined_df, pd.Timestamp.now(), norm_params)
        
        print("Running Model.Predict()");
        predictions = model.predict( [pred_input["x_features"], pred_input["x_item_idx"]])
    
        prediction_df = pred_input["prediction_df"]
        prediction_df["prediction"] = predictions
    
        dataframes = {
             "predictions": prediction_df,
             "normalized_df": normalized_df,
             "combined_df": combined_df
         }
        save_experiment(model, history, dataframes, modelBuildParams, modelTrainParams, norm_params, base_dir=baseDir)
        # return {
        #     "model": model,
        #     "history": history,
        #     "normalized_df": normalized_df,
        #     "prediction_df": prediction_df,
        #     "norm_params": norm_params
        # }
    ###############################################
    
    
    def RunPredictionsOnly(combined_df,model_dir,prediction_date):
        """
        Loads a trained model + artifacts and runs predictions only.
        """
        model = tf.keras.models.load_model(f"{model_dir}/model.keras")
    
        with open(f"{model_dir}/norm_params.json", "r") as f:
            norm_params = json.load(f)
    
        pred_input = build_prediction_input(combined_df=combined_df, prediction_date=prediction_date,norm_params=norm_params)
    
        predictions = model.predict(
            [pred_input["x_features"], pred_input["x_item_idx"]]
        )
    
        prediction_df = pred_input["prediction_df"].copy()
        prediction_df["prediction"] = predictions
    
        return prediction_df
    ###############################################

    
