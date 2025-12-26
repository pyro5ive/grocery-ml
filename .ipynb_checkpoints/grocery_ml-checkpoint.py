
import time
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import os
from datetime import datetime
import asyncio
import json

import gc
import tensorflow as tf
from tensorflow.keras import layers, models

from sklearn.model_selection import train_test_split
import numpy as np
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
from sklearn.metrics import silhouette_score

from school_features import SchoolFeatures
from weather_features import WeatherFeatures
from item_name_utils import ItemNameUtils
from temporal_features import TemporalFeatures
from holiday_features import HolidayFeatures
from wallmart_rcpt_parser import WallmartRecptParser
from winn_dixie_recpt_parser import WinnDixieRecptParser 
from hidden_layer_param_builder import HiddenLayerParamSetBuilder
asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())


class GroceryML:
    
    def __init__(self):
        pass;
        self.itemNameUtils = ItemNameUtils();
        
    ###########################################################################################
    def build_combined_df(self):
        
        winndixie_df = self.build_winn_dixie_df();
        wallmart_df = WallmartRecptParser.build_wall_mart_df("./walmart");
        #self.combined_df = pd.concat([winndixie_df, wallmart_df], ignore_index=True)
        self.combined_df = pd.concat([winndixie_df, wallmart_df[["date", "item", "source"]]],ignore_index=True)
 

        # item name and id operations
        self.canonicalize()
        self.combined_df["item"] = self.combined_df["item"].apply(ItemNameUtils.clean_item_name)
        self.combined_df = self.itemNameUtils.create_item_ids(self.combined_df)
        # 
        self.insert_negative_samples()
        
        df_weather = WeatherFeatures.BuildWeather().reset_index()
        self.combined_df = self.combined_df.merge(df_weather, on="date", how="left")

        trip_df = self.build_trip_level_features()
        self.combined_df = self.combined_df.merge(trip_df, on="date", how="left")
        self.build_purchase_item_freq_cols()

        self.build_freq_ratios()
        
        self.combined_df = TemporalFeatures.compute_days_since_last_purchase_for_item(self.combined_df)
        self.combined_df = TemporalFeatures.compute_avg_days_between_item_purchases(self.combined_df)
        self.combined_df["item_due_ratio_feat"] = self.compute_due_ratio(self.combined_df)
        
        #self.create_bulkAdjustedUrgencyRatio_for_training(self.combined_df);
        # ============================================================
        # MERGE HABIT FEATURES
        # ============================================================
        # habit_df = build_habit_features(combined_df)
        # self.combined_df = self.combined_df.merge(habit_df, on="itemId",how="left")
        # self.combined_df[["habitFrequency_feat", "habitSpan_feat", "habitDecay_feat"]] = (
        #     self.combined_df[["habitFrequency_feat", "habitSpan_feat", "habitDecay_feat"]].fillna(0.0)
        # ) 
    ###########################################################################################
    def canonicalize(self):
    
        patterns = ["prairie-farm-milk","kleinpeter-milk", "kl-milk", "Milk, Fat Free,", "Fat-Free Milk"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "milk")
        patterns = ["Bunny Bread", "White Sandwich Bread", "bunny-bread","se-grocers-bread","seg-sandwich-bread", "seg-white-bread"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "bread")
        patterns = ["dandw-cheese", "kraft-cheese", "se-grocers-cheese", "know-and-love-cheese"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "cheese")
        patterns = ["blue-plate-mayo", "blue-plate-mynnase"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "mayo")
        patterns = ["gatorade", "powerade"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "gatorade-powerade")
        patterns = ["chicken-cutlet", "chicken-leg", "chicken-thigh", "chicken-thighs"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "chicken")
        patterns = ["chobani-yogrt-flip", "chobani-yogurt"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "yogurt")
        patterns = ["coca-cola", "coca-cola-cola", "cocacola-soda", "coke", "cola"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "coke")
        patterns = ["hugbi-pies", "-hugbi-pies"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "hugbi-pies")
        patterns  = ["cereal"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "cereal")
        patterns = ["minute-maid-drink", "minute-maid-drinks", "minute-maid-lmnade"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "minute-maid-drink")
        patterns = ["egglands-best-egg", "egglands-best-eggs", "eggs"]
        self.itemNameUtils.canonicalize_items(self.combined_df, patterns, "eggs")
    ###########################################################################################

    def build_habit_features(self, df, tau_days=120):
        df = df.copy()
        df["date"] = pd.to_datetime(df["date"])
    
        total_trips = df["date"].nunique()
        timeline_days = (df["date"].max() - df["date"].min()).days or 1
    
        rows = []
    
        for itemId, g in df.groupby("itemId"):
            buys = g[g["didBuy_target"] == 1]["date"]
    
            if len(buys) == 0:
                rows.append({
                    "itemId": itemId,
                    "habitFrequency_feat": 0.0,
                    "habitSpan_feat": 0.0,
                    "habitDecay_feat": 0.0,
                })
                continue
    
            first = buys.min()
            last = buys.max()
    
            habitFrequency = len(buys) / total_trips
            habitSpan = (last - first).days / timeline_days
            days_since_last = (df["date"].max() - last).days
            habitDecay = np.exp(-days_since_last / tau_days)
    
            rows.append({
                "itemId": itemId,
                "habitFrequency_feat": habitFrequency,
                "habitSpan_feat": habitSpan,
                "habitDecay_feat": habitDecay,
            })
    
        return pd.DataFrame(rows)
    ###########################################################################################
    
    def compute_due_ratio(self, df, cap=3.0):
        ratio = df["daysSinceThisItemLastPurchased_feat"] / df["avgDaysBetweenItemPurchases_feat"]
        ratio = ratio.replace([np.inf, -np.inf], np.nan).fillna(0)
        return ratio.clip(0, cap)
    ###########################################################################################

    def compute_due_score(self, df, itemId=None, use_sigmoid=True, normalize=False, weights=None):
        
        if weights is None:
            weights = {
                "daysSinceThisItemLastPurchased_feat": 1.5,
                "freq_30_feat": 1.0,
                "freq_90_feat": 0.5
            }
    
        # --------------------------------------------------------
        # Optional itemId filter
        # --------------------------------------------------------
        if itemId is not None:
            df = df[df["itemId"] == itemId].copy()
        else:
            df = df.copy()
    
        # --------------------------------------------------------
        # RAW linear score (pre-normalization)
        # --------------------------------------------------------
        df["due_score_raw_feat"] = (
            weights["daysSinceThisItemLastPurchased_feat"] * df["daysSinceThisItemLastPurchased_feat"]
          + weights["freq_30_feat"]              * df["freq_30_feat"]
          + weights["freq_90_feat"]              * df["freq_90_feat"]
        )
    
        # --------------------------------------------------------
        # Final due_score
        # --------------------------------------------------------
        if use_sigmoid:
            df["due_score"] = 1 / (1 + np.exp(-df["due_score_raw"]))
    
        elif normalize:
            mean = df["due_score_raw"].mean()
            std  = df["due_score_raw"].std() or 1.0
            df["due_score"] = (df["due_score_raw"] - mean) / std
    
        else:
            df["due_score"] = df["due_score_raw"]
    
        return df
    ###########################################################################################

    def build_trip_level_features(self):
        print("build_trip_level_features()");
        grouped_df = ( self.combined_df[["date"]]
            .drop_duplicates()
            .sort_values("date")
            .reset_index(drop=True)
        )
    
        grouped_df["daysSinceLastTrip_feat"] = TemporalFeatures.create_days_since_last_trip(grouped_df)
        grouped_df["avgDaysBetweenTrips_feat"] = TemporalFeatures.compute_avg_days_between_trips(grouped_df)
        grouped_df["daysUntilNextHoliday_feat"]   = grouped_df["date"].apply(HolidayFeatures.compute_days_until_next_holiday)
        grouped_df["daysSinceLastHoliday_feat"]   = grouped_df["date"].apply(HolidayFeatures.compute_days_since_last_holiday)
        grouped_df["holidayProximityIndex_feat"]  = grouped_df["date"].apply(HolidayFeatures.compute_holiday_proximity_index)
        grouped_df["daysUntilSchoolStart_feat"]   = grouped_df["date"].apply(SchoolFeatures.compute_days_until_school_start)
        grouped_df["daysUntilSchoolEnd_feat"]     = grouped_df["date"].apply(SchoolFeatures.compute_days_until_school_end)
        grouped_df["schoolSeasonIndex_feat"]      = grouped_df["date"].apply(SchoolFeatures.compute_school_season_index)
    
        grouped_df = TemporalFeatures.create_date_features(grouped_df)
        return grouped_df;
    ###########################################################################
    
    def compute_bulkAdjustedUrgencyRatio_value(self, days_since, avg_between, bulk_flag, did_buy):
        """
        Compute one bulk-adjusted urgency ratio value.
        Shared by training + prediction. Caller decides did_buy.
        """
        if did_buy != 1:
            return 0.0

        bulk_factor_value = 2.5
        denominator = avg_between * (bulk_factor_value if bulk_flag == 1 else 1.0)

        if denominator == 0:
            return 0.0

        return days_since / denominator
    ###########################################################################
   
    def create_bulkAdjustedUrgencyRatio_for_training(self, df):
        """
        Create bulkAdjustedUrgencyRatio_feat using real didBuy_target values.
        """
        if "bulkFlag" not in df.columns:
            df["bulkAdjustedUrgencyRatio_feat"] = [0.0] * len(df)
            return df

        ratios = []
        for _, row in df.iterrows():
            ratios.append(self.compute_bulkAdjustedUrgencyRatio_value(
                row["daysSinceThisItemLastPurchased_feat"],
                row["avgDaysBetweenItemPurchases_feat"],
                row["bulkFlag"],
                row["didBuy_target"]      # <-- real 0/1
            ))

        df["bulkAdjustedUrgencyRatio_feat"] = ratios
        return df
    ###########################################################################
    def create_bulkAdjustedUrgencyRatio_for_prediction(self, df):
        """
        Create bulkAdjustedUrgencyRatio_feat assuming did_buy==1 for all rows.
        """
        if "bulkFlag" not in df.columns:
            df["bulkAdjustedUrgencyRatio_feat"] = [0.0] * len(df)
            return df

        ratios = []
        for _, row in df.iterrows():
            ratios.append(self.compute_bulkAdjustedUrgencyRatio_value(
                row["daysSinceThisItemLastPurchased_feat"],
                row["avgDaysBetweenItemPurchases_feat"],
                row["bulkFlag"],
                1                           # <-- always 1 at prediction time
            ))

        df["bulkAdjustedUrgencyRatio_feat"] = ratios
        return df
    ###########################################################################
    
    def build_purchase_item_freq_cols(self):

        print("build_purchase_item_freq_cols()");
        freq_windows = [7, 15, 30, 90, 365]
        max_w = max(freq_windows)
    
        # initialize output columns
        for w in freq_windows:
            self.combined_df[f"freq_{w}_feat"] = np.nan
    
        # process each itemId independently
        result_frames = []
        for item_id, group in self.combined_df.groupby("itemId", group_keys=False):
            group = group.copy()
            group = group.sort_values("date").reset_index(drop=True)
            history = []
    
            col_date = group.columns.get_loc("date")
            col_buy = group.columns.get_loc("didBuy_target")
            col_freq = {w: group.columns.get_loc(f"freq_{w}_feat") for w in freq_windows}
    
            for i in range(len(group)):
                cur_date = group.iat[i, col_date]
    
                # record a purchase occurrence
                if group.iat[i, col_buy] == 1:
                    history.append(cur_date)
    
                # prune old entries once using max window
                cutoff_max = cur_date - pd.Timedelta(days=max_w)
                history = [d for d in history if d >= cutoff_max]
    
                # compute window counts
                for w in freq_windows:
                    cutoff = cur_date - pd.Timedelta(days=w)
                    count = 0
                    for d in history:
                        if d >= cutoff:
                            count += 1
                    group.iat[i, col_freq[w]] = count
    
            result_frames.append(group)
    
        self.combined_df = pd.concat(result_frames, ignore_index=True)
    ###########################################################################################

    # def build_trip_ratio(self):
    #     self.combined_df["purchaseToTripRatio"] = combined_df["daysSinceLastPurchase"] / combined_df["avgDaysBetweenPurchases"]
    
    def build_freq_ratios(self):
        (
            self.combined_df["freq7_over30_feat"],
            self.combined_df["freq30_over365_feat"],
        ) = TemporalFeatures.compute_freq_ratios(
            self.combined_df["freq_7_feat"],
            self.combined_df["freq_30_feat"],
            self.combined_df["freq_365_feat"],
        )
    ###########################################################################################
        
    def insert_negative_samples(self):
        print("insert_negative_samples()")
    
        # keep a lookup of itemId -> item name (each itemId maps to exactly one name)
        item_lookup = (
            self.combined_df[["itemId", "item"]]
            .drop_duplicates(subset=["itemId"])
        )
    
        # 1. mark real purchases
        self.combined_df["didBuy_target"] = 1
    
        # 2. full grid
        all_items = self.combined_df["itemId"].unique()
        all_dates = self.combined_df["date"].unique()
        full = (
            pd.MultiIndex.from_product([all_dates, all_items], names=["date", "itemId"])
            .to_frame(index=False)
        )
    
        # 3. merge raw purchase rows
        df_full = full.merge(
            self.combined_df[["date", "itemId", "item", "source", "didBuy_target"]],
            on=["date", "itemId"],
            how="left"
        )
    
        # 4. fill missing didBuy
        df_full["didBuy_target"] = df_full["didBuy_target"].fillna(0).astype(int)
    
        # 5. fill missing item names using lookup
        df_full = df_full.merge(item_lookup, on="itemId", how="left", suffixes=("", "_lookup"))
        df_full["item"] = df_full["item"].fillna(df_full["item_lookup"])
        df_full = df_full.drop(columns=["item_lookup"])
    
        # 6. replace
        self.combined_df = df_full.copy()
    ###########################################################################################

    def build_winn_dixie_df(self):
        recptParser = WinnDixieRecptParser()
        rows = []
        for p in Path("winndixie rcpts/StevePhone2/pdf/text").glob("*.txt"):
            result = recptParser.parse(p.read_text(encoding="utf-8", errors="ignore"))
            for r in result["items"]:
                rows.append({
                    "source": p.name,
                    "date": result["date"],
                    "time": result["time"],
                    #"manager": result["manager"],
                    #"cashier": result["cashier"],
                    "item": r["item"]
                    #"qty": r["qty"],
                    #"reg": r["reg"],
                    #"youPay": r["youPay"],
                    #"reportedItemsSold": result["reported"],
                    #"rowsMatchReported": result["validation"]["rowsMatchReported"],
                    #"qtyMatchReported": result["validation"]["qtyMatchReported"],
                })
    
        winndixie_df = pd.DataFrame(rows)
        
        winndixie_df["date"] = pd.to_datetime(winndixie_df["date"])
        winndixie_df["time"] = winndixie_df["time"].astype(str)
        
        winndixie_df = WinnDixieRecptParser.remove_duplicate_receipt_files(winndixie_df)
        
        winndixie_df["item"] = winndixie_df["item"].str.replace(r"^know-and-love\s*", "", regex=True, case=False).str.strip()
        winndixie_df["item"] = winndixie_df["item"].str.replace(r"^seg\s*", "", regex=True, case=False).str.strip()
        winndixie_df["item"] = winndixie_df["item"].str.replace(r"^kandl\s*", "", regex=True, case=False).str.strip()

        winndixie_df = winndixie_df.sort_values(by=["date", "time"]).reset_index(drop=True)
        winndixie_df = winndixie_df.drop(columns=["time"])
        return winndixie_df;
    ###########################################################################################

    def export_df_to_excel_table(self, df, base_path_without_ext, sheet_name="Data"):
        """
        Export a pandas DataFrame to an Excel file as a proper Excel Table
        with no duplicated header rows.
        """
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo

        file_path = f"{base_path_without_ext}.xlsx"
        print(f"Writing XLSX: {file_path}")

        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
    
        end_row = worksheet.max_row
        end_col = worksheet.max_column
        end_col_letter = worksheet.cell(row=1, column=end_col).column_letter
    
        table_ref = f"A1:{end_col_letter}{end_row}"
        table = Table(displayName="DataTable", ref=table_ref)
    
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    
        table.tableStyleInfo = style
        worksheet.add_table(table)
    
        workbook.save(file_path)
        print(f"   XLSX Done: {file_path}")
    ###########################################################################################    

    def export_dataframe_to_csv(self, df, base_path_without_ext):
        file_path = f"{base_path_without_ext}.csv"
        print(f"Writing CSV: {file_path}")
        df.to_csv(file_path, index=True)
        print(f"  CSV done: {file_path}")
    ###########################################################################################    

    def export_dataframes_with_exp_name(self, dataframes, path, exp_suffix):
        for name, df in dataframes.items():
            base = os.path.join(path, f"{name}-{exp_suffix}")
            self.export_df_to_excel_table(df, base, sheet_name=f"{name}")
            #self.export_dataframe_to_csv(df, base)
    ###########################################################################################
    def write_json(self, obj, path):
        f = open(path, "w")
        json.dump(obj, f, indent=2)
        f.close()
    ###########################################################################################
        
    def save_experiment(self,model, history, dataframes, build_params, train_params, norm_params, base_dir):
        name_parts = []
       
        if "embedding_dim" in build_params:
            name_parts.append(f"e{build_params['embedding_dim']}")
        if "layers" in build_params:
            hl = "-".join(str(x) for x in build_params["layers"])
            name_parts.append(f"l{hl}")
        if "epochs" in train_params:
            name_parts.append(f"ep{train_params['epochs']}")
        if "output_activation" in build_params:
            name_parts.append(f"oa_{build_params['output_activation']}")

        # add uniqueness to avoid overwrite
        base_name = "__".join(name_parts) if name_parts else "exp_unlabeled"
        short_id = str(abs(hash(time.time())))[:6]
        exp_name = f"{base_name}__{short_id}"

        exp_dir = os.path.join(base_dir, exp_name)      
        
        print(f"Creating dir: {exp_dir}");
        os.makedirs(exp_dir, exist_ok=True);
        
        print("Exporting dataframes:");
        self.export_dataframes_with_exp_name(dataframes, exp_dir, exp_name)

        modelDir = os.path.join(exp_dir, "model")
        print(f"Creating model dir: {modelDir}");
        os.makedirs(modelDir, exist_ok=True)
        
        print("Saving Model")
        model.save(modelDir)
        model.save_weights(os.path.join(modelDir, "weights.h5"))
        self.write_json(history.history, os.path.join(modelDir, "history.json"))
        self.write_json(build_params,    os.path.join(exp_dir,  "build_params.json"))
        self.write_json(train_params,    os.path.join(exp_dir,  "train_params.json"))
        self.write_json(norm_params,     os.path.join(exp_dir,  "norm_params.json"))
    
        print("Saved experiment →", exp_dir)
    ###########################################################################################
    
    def fit_normalization_params(self,combined_df):
        params = {}
        feature_cols = [c for c in combined_df.columns if c.endswith("_feat")]
        cyc_cols = [c for c in feature_cols if c.endswith("_cyc_feat")]
        num_cols = [c for c in feature_cols if c not in cyc_cols]
        
        for col in num_cols:
            params[col] = {"mean": combined_df[col].mean(),"std": combined_df[col].std()}
    
        for col in cyc_cols:
            params[col] = {"period": TemporalFeatures.get_period_for_column(col)}
    
        return params
    ###########################################################################################
    
    def normalize_features(self,combined_df, norm_params):
    
        print("normalize_features()");
        normalized_df = combined_df.copy()
        for col, cfg in norm_params.items():
            if col.endswith("_cyc_feat"):
                sin_col, cos_col = TemporalFeatures.encode_sin_cos(combined_df[col], cfg["period"])
                normalized_df[f"{col}_sin_norm"] = sin_col
                normalized_df[f"{col}_cos_norm"] = cos_col
                normalized_df.drop(columns=[col], inplace=True)
    
            else:
                mean_val = cfg["mean"]
                std_val = cfg["std"]
                norm_col = col.replace("_feat", "_norm")
    
                if std_val == 0:
                    normalized_df[norm_col] = 0.0
                else:
                    normalized_df[norm_col] = (combined_df[col] - mean_val) / std_val
    
                normalized_df.drop(columns=[col], inplace=True)
    
        return normalized_df
    ###########################################################################################    

    def build_prediction_input(self, combined_df, prediction_date, norm_params):
    
        print("build_prediction_input()")
        print(f"Prediction date: {prediction_date.strftime('%Y-%m-%d')}")
    
        latest_rows_df = (
            combined_df.sort_values("date")
            .groupby("itemId")
            .tail(1)
            .copy()
            .reset_index(drop=True)
        )
    
        latest_rows_df["date"] = prediction_date
    
        # days since last trip
        max_data_date = combined_df["date"].max()
        days_forward = (prediction_date - max_data_date).days
        latest_rows_df["daysSinceLastTrip_feat"] = days_forward
    
        # avg days between trips (global)
        latest_rows_df["avgDaysBetweenTrips_feat"] = combined_df["avgDaysBetweenTrips_feat"].iloc[-1]
    
        # === extend "days since this item last purchased"
        # last known value already correct per item at max_data_date
        last_vals = combined_df.sort_values("date").groupby("itemId").tail(1)
        latest_rows_df["daysSinceThisItemLastPurchased_feat"] = (
            last_vals["daysSinceThisItemLastPurchased_feat"].values + days_forward
        )
    
        # === extend avgDaysBetweenItemPurchases per item
        last_avg_vals = combined_df.sort_values("date").groupby("itemId").tail(1)
        latest_rows_df["avgDaysBetweenItemPurchases_feat"] = last_avg_vals["avgDaysBetweenItemPurchases_feat"].values
    
        latest_rows_df[[
            "daysSinceThisItemLastPurchased_feat",
            "avgDaysBetweenItemPurchases_feat"
        ]] = latest_rows_df[[
            "daysSinceThisItemLastPurchased_feat",
            "avgDaysBetweenItemPurchases_feat"
        ]].fillna(0)
    
        latest_rows_df["item_due_ratio_feat"] = TemporalFeatures.compute_due_ratio(latest_rows_df)
    
        latest_rows_df["daysUntilNextHoliday_feat"] = HolidayFeatures.compute_days_until_next_holiday(prediction_date)
        latest_rows_df["daysSinceLastHoliday_feat"] = HolidayFeatures.compute_days_since_last_holiday(prediction_date)
        latest_rows_df["holidayProximityIndex_feat"] = HolidayFeatures.compute_holiday_proximity_index(prediction_date)
        latest_rows_df["daysUntilSchoolStart_feat"] = SchoolFeatures.compute_days_until_school_start(prediction_date)
        latest_rows_df["daysUntilSchoolEnd_feat"] = SchoolFeatures.compute_days_until_school_end(prediction_date)
        latest_rows_df["schoolSeasonIndex_feat"] = SchoolFeatures.compute_school_season_index(prediction_date)
    
        latest_rows_df["year_feat"] = prediction_date.year
        latest_rows_df["month_cyc_feat"] = prediction_date.month
        latest_rows_df["day_cyc_feat"] = prediction_date.day
        latest_rows_df["dow_cyc_feat"] = prediction_date.weekday()
        latest_rows_df["doy_feat"] = prediction_date.timetuple().tm_yday
        latest_rows_df["quarter_feat"] = ((prediction_date.month - 1) // 3) + 1
    
        if "didBuy_target" in latest_rows_df.columns:
            latest_rows_df.drop(columns=["didBuy_target"], inplace=True)
    
        normalized_latest_rows_df = self.normalize_features(latest_rows_df, norm_params)
    
        feature_cols = [c for c in normalized_latest_rows_df.columns if c.endswith("_norm")]
        x_features = normalized_latest_rows_df[feature_cols].to_numpy(np.float32)
        x_item_idx = normalized_latest_rows_df["itemId"].to_numpy(np.int32)
    
        self.export_df_to_excel_table(normalized_latest_rows_df, "normalized_latest_rows_df.xlsx", ".")
        self.export_df_to_excel_table(latest_rows_df, "latest_rows_df.xlsx", ".")
    
        return {
            "prediction_df": normalized_latest_rows_df,
            "x_features": x_features,
            "x_item_idx": x_item_idx,
            "feature_cols": feature_cols
        }
    ###########################################################################################



    def build_and_compile_model(self,feat_cols_count, item_count, build_params):
        num_in = layers.Input(shape=(feat_cols_count,))
        item_in = layers.Input(shape=(), dtype="int32")
    
        emb = layers.Embedding(
            input_dim=item_count,
            output_dim=build_params["embedding_dim"]
        )(item_in)
    
        x = layers.Concatenate()([num_in, layers.Flatten()(emb)])
    
        for neuron_count in build_params["layers"]:
            x = layers.Dense(neuron_count, activation=build_params["activation"])(x)
    
        out = layers.Dense(1, activation=build_params["output_activation"])(x)
    
        model = models.Model(inputs=[num_in, item_in], outputs=out)
    
        optimizer_name = build_params.get("optimizer", "adam")
        learning_rate = build_params.get("learning_rate")
    
        if optimizer_name == "adam":
            optimizer = tf.keras.optimizers.Adam(learning_rate=learning_rate)
        elif optimizer_name == "adamw":
            optimizer = tf.keras.optimizers.AdamW(learning_rate=learning_rate)
        else:
            raise ValueError(f"Unsupported optimizer: {optimizer_name}")
    
        model.compile(
            optimizer=optimizer,
            loss=build_params.get("loss", "mse"),
            metrics=build_params.get("metrics", ["mae"])
        )
    
        return model
    ###########################################################################################
    
    def train_model(self, model, df, feature_cols, target_col, train_params):
        print("train_model()");
        x_feat = df[feature_cols].to_numpy(np.float32)
        x_item = df["itemId"].to_numpy(np.int32)
        y = df[target_col].to_numpy(np.float32)
    
        x_feat_tr, x_feat_te, x_item_tr, x_item_te, y_tr, y_te = train_test_split(
            x_feat, x_item, y, test_size=0.2, random_state=42
        )
    
        history = model.fit(
            [x_feat_tr, x_item_tr],
            y_tr,
            validation_split=0.1,
            epochs=train_params["epochs"],
            batch_size=train_params.get("batch_size", 32),
            verbose=0
        )
    
        return history
    ###########################################################################################
  
    def run_experiment(self, combined_df, modelBuildParams, modelTrainParams, baseDir):
        
        print("run_experiment() ");
        norm_params = self.fit_normalization_params(combined_df)
        normalized_df = self.normalize_features(combined_df, norm_params)
    
        feature_cols = [c for c in normalized_df.columns if c.endswith("_norm")]
        target_cols = [c for c in normalized_df.columns if c.endswith("_target")]
    
        if len(target_cols) != 1:
            raise ValueError("Exactly one target column is required")
        target_col = target_cols[0]
    
        feat_cols_count = len(feature_cols)
        item_count = int(normalized_df["itemId"].max()) + 1
    
        model = self.build_and_compile_model(feat_cols_count, item_count, modelBuildParams)
    
        history = self.train_model(model, normalized_df, feature_cols, target_col, modelTrainParams)
       
        # pred_input = build_prediction_input_df(combined_df, normalized_df["date"].max(), norm_params)
        pred_input = self.build_prediction_input(combined_df, pd.Timestamp.now(), norm_params)
        
        print("Running Model.Predict()");
        predictions = model.predict( [pred_input["x_features"], pred_input["x_item_idx"]])
    
        prediction_df = pred_input["prediction_df"]
        prediction_df.insert(3, "prediction",  predictions)    
        prediction_df = self.itemNameUtils.map_item_ids_to_names(prediction_df)
        prediction_df = prediction_df.sort_values("prediction", ascending=False).reset_index(drop=True)
        
        dataframes = {
            "predictions": prediction_df,
            "normalized_df": normalized_df,
            "combined_df": combined_df
        }
        self.save_experiment(model, history, dataframes, modelBuildParams, modelTrainParams, norm_params, base_dir=baseDir)
    ###########################################################################################
    def RunPredictionsOnly(self, combined_df, model_dir, prediction_date):
        """
        Loads a trained model + artifacts and runs predictions only.
        Same behavior as run_experiment(), minus training + saving.
        """
        print("RunPredictionsOnly()")
    
        # load model + artifacts
        model = tf.keras.models.load_model(model_dir)
    
        with open(os.path.join(model_dir, "norm_params.json"), "r") as f:
            norm_params = json.load(f)
    
        # build input rows
        pred_input = self.build_prediction_input(combined_df, prediction_date, norm_params)
    
        # run predict
        print("Running Model.Predict()")
        predictions = model.predict([pred_input["x_features"], pred_input["x_item_idx"]])
    
        # attach + post-process
        prediction_df = pred_input["prediction_df"].copy()
        prediction_df["prediction"] = predictions
        prediction_df = self.itemNameUtils.map_item_ids_to_names(prediction_df)
        prediction_df = prediction_df.sort_values("prediction", ascending=False).reset_index(drop=True)
    
        return prediction_df
    ###########################################################################################
   
