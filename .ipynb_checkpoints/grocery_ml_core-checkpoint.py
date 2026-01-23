import time
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import asyncio
import json
import gc
from sklearn.model_selection import train_test_split
import numpy as np
from sklearn.preprocessing import LabelEncoder
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt
from sklearn.metrics import silhouette_score
import logging
#
from school_features import SchoolFeatures
from weather_features import WeatherFeatures
from item_name_utils import ItemNameUtils
from temporal_features_2 import TemporalFeatures
from data_creator import DataCreator
from holiday_features import HolidayFeatures
from wallmart_rcpt_parser import WallmartRecptParser
from winn_dixie_recpt_parser import WinnDixieRecptParser 
from hidden_layer_param_builder import HiddenLayerParamSetBuilder
from weather_service import NwsWeatherService;
from usda import UsdaCategoryEncoder
from usda import UsdaFoodDataService
import item_name_constants


logger = logging.getLogger(__name__)

class GroceryMLCore:

    usdaCatEncoder = None;
    usdaApiService = None;
    itemNameUtils = None; 
    weatherService = None;
    
 
    def __init__(self):
        pass;
        self.itemNameUtils = ItemNameUtils();
        self.weatherService = NwsWeatherService();
        self.usdaApiService =  UsdaFoodDataService();
        self.usdaCatEncoder = UsdaCategoryEncoder(self.usdaApiService);
   ###########################################################################################
    def validate_no_empty_columns(self, df, exclude_cols=None):
        print("validate_no_empty_columns()")
        if exclude_cols is None:
            exclude_cols = []
        #
        bad_cols = [
            c for c in df.columns
            if c not in exclude_cols and df[c].isna().any()
        ]
    
        if bad_cols:
            raise ValueError(f"Columns contain empty values: {bad_cols}")
  ###########################################################################################
    def drop_rare_purchases(self, df):
        print("drop_rare_purchases()")
        df = df[df["itemPurchaseCount_raw"] != 1].reset_index(drop=True)
        return df;
    ###########################################################################################
    def log_feature(self, values: pd.Series) -> pd.Series:
        # guard: negatives are clipped, zeros allowed
        values = values.clip(lower=0)
        return np.log1p(values)
    ###########################################################################################
    def normalize_item_names(self, df):
        df = self.itemNameUtils.remove_items_matching_terms(df, "item", item_name_constants.EXCLUDED_ITEMS);
        # elf._combined_df["itemName_lemma"] = self._combined_df["item"].apply(self.lemmatize_item_name)
        df = self.itemNameUtils.strip_prefixes_from_column(df ,"item", item_name_constants.BRAND_PREFIXES);
        df["item"] = df["item"].apply(self.itemNameUtils.clean_item_name)
        df = self.itemNameUtils.canonicalize(df)
        return df
    ###########################################################################################
    def get_feature_col_names(self, df):
        """Returns feature columns before normalization."""
        return [
            c for c in df.columns
            if c.endswith("_feat") or c.endswith("_cyc_feat")
        ]
    ###########################################################################################
    def build_trip_interveral_feautres(self, df):
        print("build_trip_interveral_feautres(): start")
        trip_df = (df[["date"]] .drop_duplicates() .sort_values("date") .reset_index(drop=True))
        trip_df["daysSinceLastTrip_raw"] = TemporalFeatures.create_days_since_last_trip(trip_df)
        trip_df["avgDaysBetweenTrips_feat"] = TemporalFeatures.compute_avg_days_between_trips(trip_df)
        print("build_trip_interveral_feautres(): done")
        return df.merge(trip_df, on="date", how="left")
    ##############################################################################################
    def create_item_supply_level_feat(self, df):
        print("create_item_supply_level_feat()")
    
        try:
            ratio = np.where(
                df["avgDaysBetweenItemPurchases_feat"] > 0,
                df["daysSinceThisItemLastPurchased_raw"] / df["avgDaysBetweenItemPurchases_feat"],
                0.0
            )
    
            df["itemSupplyLevel_feat"] = np.clip(1.0 - ratio, 0.0, 1.0)
    
        except Exception as ex:
            print("create_item_supply_level_feat() failed")
            print(ex)
            raise
    
        return df
    ###########################################################################################
    def create_didBuy_target_col(self, df, colName):
        print(f"creating target col: {colName}");
        df[colName] = 1
        return df; 
    ##############################################################################################
    def add_item_total_purchase_count_feat(self, df, feature_name: str):
        """
        Adds a history-only cumulative purchase count per item.
        For each row, the count reflects how many times the item
        has been purchased up to and including that day.
        """
    
        print("add_item_total_purchase_count_feat()")
    
        # Ensure correct temporal order per item
        df = df.sort_values(["itemId", "date"]).copy()
    
        # History-only cumulative count
        df[feature_name] = (df.groupby("itemId")["didBuy_target"] .cumsum() .astype(int) )
    
        return df
    ##############################################################################################
    def build_holiday_features(self, df):
        print("build_holiday_features()")
    
        df = df.drop(columns=[c for c in df.columns if c.endswith("_holiday_feat")], errors="ignore")
    
        grouped_df = (
            df[["date"]]
            .drop_duplicates()
            .sort_values("date")
            .reset_index(drop=True)
        )
    
        holiday_feats = HolidayFeatures.build_federal_holiday_flag_and_proximity_features(grouped_df["date"])
        holiday_feats["date"] = grouped_df["date"].values
    
        df = df.merge(holiday_feats, on="date", how="left")
        return df
    ##############################################################################################
    #  def build_holiday_features(self, df):
   #      print("build_holiday_features()")
   #      df = df.drop(columns=[c for c in df.columns if c.endswith("_holiday_feat")], errors="ignore")
   #      grouped_df = (
   #          df[["date"]]
   #          .drop_duplicates()
   #          .sort_values("date")
   #          .reset_index(drop=True)
   #      )

   #      grouped_df = HolidayFeatures.build_federal_holiday_flag_and_proximity_features(grouped_df["date"])
        
   #      # grouped_df["daysUntilNextHoliday_raw"] = HolidayFeatures.compute_days_until_next_holiday(grouped_df["date"])
   #      # grouped_df["daysSinceLastHoliday_raw"] = HolidayFeatures.compute_days_since_last_holiday(grouped_df["date"])
   #      # grouped_df["holidayProximity_feat"] = HolidayFeatures.compute_holiday_proximity_index(grouped_df["date"])
    
   #      df = df.merge(grouped_df, on="date", how="left")
   #      return df
   # ##############################################################################################

    def build_school_schedule_features(self, df):
        print("build_school_schedule_features(): start")
        
        df = df.drop(columns=["daysUntilSchoolStart_raw","daysUntilSchoolEnd_raw","schoolSeasonIndex_feat"], errors="ignore")
        grouped_df = (df[["date"]].drop_duplicates().sort_values("date").reset_index(drop=True))
        grouped_df["daysUntilSchoolStart_raw"] = SchoolFeatures.compute_days_until_school_start(grouped_df["date"])
        grouped_df["daysUntilSchoolEnd_raw"] = SchoolFeatures.compute_days_until_school_end(grouped_df["date"])
        grouped_df["schoolSeasonIndex_feat"] = SchoolFeatures.compute_school_season_index(grouped_df["date"])
        df = df.merge(grouped_df, on="date", how="left")
        print("build_school_schedule_features(): done")
        return df
    ##############################################################################################
    ## TODO:build_purchase_item_freq_cols  is broken 
    # def build_purchase_item_freq_cols(self, df):

    #     print("build_purchase_item_freq_cols()");
    #     freq_windows = [7, 15, 30, 90, 365]
    #     max_w = max(freq_windows)
    
    #     # initialize output columns
    #     for w in freq_windows:
    #         df[f"freq_{w}_feat"] = np.nan
    
    #     # process each itemId independently
    #     result_frames = []
    #     for item_id, group in df.groupby("itemId", group_keys=False):
    #         group = group.copy()
    #         group = group.sort_values("date").reset_index(drop=True)
    #         history = []
    
    #         col_date = group.columns.get_loc("date")
    #         col_buy = group.columns.get_loc("didBuy_target")
    #         col_freq = {w: group.columns.get_loc(f"freq_{w}_feat") for w in freq_windows}
    
    #         for i in range(len(group)):
    #             cur_date = group.iat[i, col_date]
    
    #             # record a purchase occurrence
    #             if group.iat[i, col_buy] == 1:
    #                 history.append(cur_date)
    
    #             # prune old entries once using max window
    #             cutoff_max = cur_date - pd.Timedelta(days=max_w)
    #             history = [d for d in history if d >= cutoff_max]
    
    #             # compute window counts
    #             for w in freq_windows:
    #                 cutoff = cur_date - pd.Timedelta(days=w)
    #                 count = 0
    #                 for d in history:
    #                     if d >= cutoff:
    #                         count += 1
    #                 group.iat[i, col_freq[w]] = count
    
    #         result_frames.append(group)
    
    #     return pd.concat(result_frames, ignore_index=True)
    ###########################################################################################    
    def insert_negative_samples(self, df):
        print("insert_negative_samples()")
    
        # ensure purchase flag exists
        df = df.copy()
    
        # itemId → item name lookup
        item_lookup = (
            df[["itemId", "item"]]
            .drop_duplicates(subset=["itemId"])
        )
    
        # first purchase date per item (activation point)
        first_purchase = (
            df[df["didBuy_target"] == 1]
            .groupby("itemId")["date"]
            .min()
        )
    
        # build valid (date, itemId) pairs ONLY after activation
        rows = []
        all_dates = df["date"].unique()
    
        for itemId, first_date in first_purchase.items():
            valid_dates = all_dates[all_dates >= first_date]
            for d in valid_dates:
                rows.append({"date": d, "itemId": itemId})
    
        full = pd.DataFrame(rows)
    
        # merge back original data
        df_full = full.merge(df, on=["date", "itemId"], how="left")
    
        # fill negatives
        df_full["didBuy_target"] = df_full["didBuy_target"].fillna(0).astype(int)
    
        # restore item names
        df_full = df_full.merge(item_lookup, on="itemId", how="left", suffixes=("", "_lookup"))
        df_full["item"] = df_full["item"].fillna(df_full["item_lookup"])
        df_full = df_full.drop(columns=["item_lookup"])
    
        # fill source fields for negatives
        df_full["source"] = df_full["source"].fillna("_neg_sample_").astype(str)
        
        return df_full
    ###########################################################################################            
    def create_full_calendar_and_merge(self, df, days: int = 365) -> pd.DataFrame:
        df = df.copy()
        df["date"] = pd.to_datetime(df["date"]).dt.normalize()
    
        # lookup that already exists in source df
        item_lookup = df[["itemId", "item"]].drop_duplicates("itemId")
    
        max_date = df["date"].max()
        min_date = max_date - pd.Timedelta(days=days - 1)
    
        calendar = (
            item_lookup[["itemId"]]
            .merge(
                pd.DataFrame({"date": pd.date_range(min_date, max_date, freq="D")}),
                how="cross"
            )
        )
    
        merged = calendar.merge(df, on=["itemId", "date"], how="left")
    
        # fill required fields
        merged["didBuy_target"] = merged["didBuy_target"].fillna(0).astype(int)
        merged["source"] = merged["source"].fillna("_neg_sample_")
    
        # restore item deterministically (no NaNs possible)
        merged = merged.merge(item_lookup, on="itemId", how="left", suffixes=("", "_lk"))
        merged["item"] = merged["item"].fillna(merged["item_lk"])
        merged = merged.drop(columns=["item_lk"])
    
        merged = merged.sort_values(["itemId", "date"]).reset_index(drop=True)
    
        return merged[["date", "source", "itemId", "item", "qty", "didBuy_target"]]
    ############################################################################################
    def build_winn_dixie_additional_text_rcpts_df(self, folderPath):
        recptParser = WinnDixieRecptParser()
        rows = []
        for p in Path(folderPath).glob("*.txt"):
            result = recptParser.parse(p.read_text(encoding="utf-8", errors="ignore"))
            for r in result["items"]:
                rows.append({
                    "source": p.name,
                    "date": result["date"],
                    "time": result["time"],
                    #"manager": result["manager"],
                    #"cashier_raw": result["cashier"],
                    "item": r["item"],
                    "qty": r["qty"],
                    #"reg": r["reg"],
                    #"youPay": r["youPay"],
                    #"reportedItemsSold": result["reported"],
                    #"rowsMatchReported": result["validation"]["rowsMatchReported"],
                    #"qtyMatchReported": result["validation"]["qtyMatchReported"],
                })
    
        additional_rcpts_df = pd.DataFrame(rows)
        
        additional_rcpts_df["date"] = pd.to_datetime(additional_rcpts_df["date"])
        additional_rcpts_df["time"] = additional_rcpts_df["time"].astype(str)
        
        additional_rcpts_df = WinnDixieRecptParser.remove_duplicate_receipt_files(additional_rcpts_df)
        
        # #additional_rcpts_df["item"] = additional_rcpts_df["item"].str.replace(r"^know-and-love\s*", "", regex=True, case=False).str.strip()
        # additional_rcpts_df["item"] = additional_rcpts_df["item"].str.replace(r"^seg\s*", "", regex=True, case=False).str.strip()
        # additional_rcpts_df["item"] = additional_rcpts_df["item"].str.replace(r"^kandl\s*", "", regex=True, case=False).str.strip()

        additional_rcpts_df = additional_rcpts_df.sort_values(by=["date", "time"]).reset_index(drop=True)
        additional_rcpts_df = additional_rcpts_df.drop(columns=["time"])
        return additional_rcpts_df;
    ###########################################################################################
    def build_winn_dixie_df(self, path):
        recptParser = WinnDixieRecptParser()
        rows = []
        for p in Path(path).glob("*.txt"):
            result = recptParser.parse(p.read_text(encoding="utf-8", errors="ignore"))
            for r in result["items"]:
                rows.append({
                    "source": p.name,
                    "date": result["date"],
                    "time": result["time"],
                    #"manager": result["manager"],
                    #"cashier_raw": result["cashier"],
                    "item": r["item"],
                    "qty": r["qty"],
                    #"reg": r["reg"],
                    #"youPay": r["youPay"],
                    #"reportedItemsSold": result["reported"],
                    #"rowsMatchReported": result["validation"]["rowsMatchReported"],
                    #"qtyMatchReported": result["validation"]["qtyMatchReported"],
                })
    
        winndixie_df = pd.DataFrame(rows)
        
        winndixie_df["date"] = pd.to_datetime(winndixie_df["date"])
        winndixie_df["time"] = winndixie_df["time"].astype(str)
        
        winndixie_df = WinnDixieRecptParser.remove_duplicate_receipt_files(winndixie_df)
        
        # winndixie_df["item"] = winndixie_df["item"].str.replace(r"^know-and-love\s*", "", regex=True, case=False).str.strip()
        # winndixie_df["item"] = winndixie_df["item"].str.replace(r"^seg\s*", "", regex=True, case=False).str.strip()
        # winndixie_df["item"] = winndixie_df["item"].str.replace(r"^kandl\s*", "", regex=True, case=False).str.strip()

        # winndixie_df["source"] = "winndixie-{";
        winndixie_df = winndixie_df.sort_values(by=["date", "time"]).reset_index(drop=True)
        winndixie_df = winndixie_df.drop(columns=["time"])
        return winndixie_df;
    ###########################################################################################
    def export_df_to_excel_table(self, df, base_path_without_ext, sheet_name="Data"):
        """
        Export a pandas DataFrame to an Excel file as a proper Excel Table
        with no duplicated header rows.
        """
        from openpyxl import load_workbook
        from openpyxl.worksheet.table import Table, TableStyleInfo

        file_path = f"{base_path_without_ext}.xlsx"
        print(f"Writing XLSX: {file_path}")

        df.to_excel(file_path, sheet_name=sheet_name, index=False)
    
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
    
        end_row = worksheet.max_row
        end_col = worksheet.max_column
        end_col_letter = worksheet.cell(row=1, column=end_col).column_letter
    
        table_ref = f"A1:{end_col_letter}{end_row}"
        table = Table(displayName="DataTable", ref=table_ref)
    
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
    
        table.tableStyleInfo = style
        worksheet.add_table(table)
    
        workbook.save(file_path)
        print(f"   XLSX Done: {file_path}")
    ###########################################################################################    
    def export_dataframe_to_csv(self, df, base_path_without_ext):
        file_path = f"{base_path_without_ext}.csv"
        logger.info("Exporting DF to CSV  path=%s", file_path)
        df.to_csv(file_path, index=True)
        logger.info("Exporting DF to CSV Complete  path=%s", file_path)
    ###########################################################################################    
    def export_dataframe_to_parquet(self, df, base_path_without_ext):
        file_path = f"{base_path_without_ext}.parquet"
        print(f"Writing PARQUET: {file_path}")
        df.to_parquet(file_path, index=True)
        print(f"  PARQUET done: {file_path}")
    ###########################################################################################
    def export_dataframes_with_exp_name(self, dataframes, path):
        for name, df in dataframes.items():
            base = os.path.join(path, f"{name}")
            self.export_df_to_excel_table(df, base, sheet_name=f"{name}")
            #self.export_dataframe_to_csv(df, base)
    ###########################################################################################    
    def write_json(self, obj, path):
        f = open(path, "w")
        json.dump(obj, f, indent=2)
        f.close()  
    ###########################################################################################    
